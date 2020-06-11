# Initialize
    from xlwt import Workbook, load_workbook
    
    wb = Workbook()                     # Create new workbook
    wb.save(filename)
    
    wb = load_workbook(filename)        # Load existing workbook

# Sheet write
    
    import xlwt 
    from xlwt import Workbook 
    
    wb = Workbook() 
    sheet1 = wb.add_sheet('Sheet 1') 
    sheet1.write(1, 0, 'ISBT DEHRADUN')  # 1,0 means position
    sheet1.write(0, 1, 'ISBT DEHRADUN') 
    wb.save('xlwt example.xls') 

    # select sheet and remove
    ref = wb['Sheet 1']
    wb.remove(ref)

# Sheet read
    import xlrd 
    data = xlrd.open_workbook('irl2.xls')               # 打开Excel文件
    table = data.sheet_by_index(0)                      # 获取第一个工作表
    nrows = table.nrows                                 # 获取行数
    ncols = table.ncols                                 # 获取列数
    excel_list = []                                     # 定义excel_list
    for row in range(1, nrows):
        for col in range(ncols):                        
            cell_value = table.cell(row, col).value     # 获取单元格数据
            excel_list.append(cell_value)               # 把数据追加到excel_list中

# write data from panda
    from openpyxl.drawing.image import Image
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Activate worksheet to write dataframe
    active = wb['sheet1']

    # Write dataframe to active worksheet
    for x in dataframe_to_rows(df):
        active.append(x)
        
    # Save workbook to write
    wb.save(filepath)

# image to excel
    import matplotlib.pyplot as plt
    # Create dataset
    x = df['dates']
    y = df['values']

    # Create and save plot as png
    plt.plot(x,y)
    plt.savefig(filepath)

    # Activate worksheet
    active = wb['sheet1']

    # Insert plot into worksheet
    # Select active sheet and cell reference
    img = Image(filepath)
    active.add_image(img,'A1')

    # Save workbook 
    wb.save(filepath)

# Write list into excel
    import pandas as pd
    df = pd.DataFrame(list)
    writer = pd.ExcelWriter('new.xlsx', engine='xlsxwriter')
    df.to_excel(writer,sheet_name='welcome',index=False)
    writer.save()
